from openpyxl import Workbook

# Reference: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook

# 워크북은 항상 최소한 하나의 워크시트를 갖고 있음
wb = Workbook()

# 활성화된 워크시트 가져 오기
# openpyxl.workbook.Workbook.active()
ws = wb.active

# 새로운 워크시트 생성
ws1 = wb.create_sheet("My sheet1")      # 가장 마지막 위치에 추가(기본값)
ws2 = wb.create_sheet("My sheet2", 0)   # 가장 앞에 추가

# 워크시트 제목 변경
ws.title = "New Title"

# 워크시트 제목 탭 색상 변경
ws.sheet_properties.tabColor = "1072BA"     # RRGGBB

# 워크시트 제목을 설정 했으면 설정한 제목을 사용해서 워크시트를 가져올 수 있음
ws3 = wb["New Title"]

# 워크북에 있는 워크시트 이름 모두 가져오기 (openpyxl.workbook.Workbook.sheetnames())
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

# 데이터 사용하기
# 하나의 셀 데이터 사용하기
c = ws['A4']

# 셀 데이터 생성하기
ws['A4'] = 4

# 행과 열을 사용해서 셀에 접근하기 openpyxl.worksheet.Worksheet.cell() 메서드 사용
d = ws.cell(row=4, column=2, value=10)

# 여러 셀 데이터 접근하기
cell_range = ws['A1':'C2']

# 행과 열 데이터 사용하기
col_c = ws['C']
col_range = ws['C:D']
row_10 = ws[10]
row_range = ws[5:10]

# openpyxl.worksheet.Worksheet.iter_rows() 메서드 사용하기
print('access row cell')
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# 열 데이터 접근
print('access col cell')
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

# 파일의 모든 열과 행이 셀 탐색하기
# openpyxl.worksheet.Worksheet.rows()
print('rows() method')
ws = wb.active
ws['C9'] = 'hello world'
print(tuple(ws.rows))

# openpyxl.worksheet.Worksheet.columns()
print('columns() method')
print(tuple(ws.columns))

# 데이터 저장
# openpyxl.cell.Cell 객체를 갖고 있으면 값을 할당할 수 있음
c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)

# 엑셀 파일 저장
wb.save('sample.xlsx')

# 엑셀 파일 읽기
from openpyxl import load_workbook
