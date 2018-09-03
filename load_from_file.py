from openpyxl import load_workbook
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)
